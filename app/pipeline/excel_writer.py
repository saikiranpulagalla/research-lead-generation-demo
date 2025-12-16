"""
Excel Writer Module
Converts JSON data to formatted Excel files with lead generation support.
"""

"""
Excel Writer Module
Converts JSON data to formatted Excel files with lead generation support.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List
from io import BytesIO


class ExcelWriter:
    """Converts structured JSON to Excel format."""
    
    def __init__(self):
        """Initialize Excel writer with formatting options."""
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.score_fill_high = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.score_font_high = Font(color="FFFFFF", bold=True)
    
    def json_to_excel(self, data: Dict[str, Any], output_path: str = None) -> BytesIO:
        """
        Convert JSON data to Excel file with flattened structure or lead generation format.
        
        Args:
            data: Extracted structured data
            output_path: Optional path to save Excel file
            
        Returns:
            BytesIO object containing Excel file
        """
        # Check if this is lead generation profile data
        if self._is_lead_gen_data(data):
            return self._profiles_to_excel(data, output_path)
        else:
            return self._flatten_to_excel(data, output_path)
    
    def _is_lead_gen_data(self, data: Dict[str, Any]) -> bool:
        """Check if data is lead generation profiles (contains probability_score)."""
        for section_name, section_data in data.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                first_item = section_data[0]
                if isinstance(first_item, dict) and "probability_score" in first_item:
                    return True
        return False
    
    def _profiles_to_excel(self, data: Dict[str, Any], output_path: str = None) -> BytesIO:
        """
        Convert lead generation profiles to Excel format.
        
        Args:
            data: Dictionary with profiles list
            output_path: Optional path to save Excel file
            
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Research Profiles"
        
        # Extract profiles from data
        profiles = []
        for section_name, section_data in data.items():
            if isinstance(section_data, list):
                profiles = section_data
                break
        
        # Sort by probability_score descending
        profiles = sorted(
            profiles,
            key=lambda x: x.get("probability_score", 0),
            reverse=True
        )
        
        # Write header row with lead gen columns
        header = [
            "Rank", "Score", "Name", "Title", "Company", "Location",
            "Email", "LinkedIn", "Keywords", "Year"
        ]
        ws.append(header)
        self._format_header_row(ws, 1)
        
        # Write profile rows
        for rank, profile in enumerate(profiles, 1):
            row_data = [
                rank,
                profile.get("probability_score", 0),
                profile.get("author_name", "N/A"),
                profile.get("title", "N/A"),
                profile.get("affiliation", "N/A"),
                profile.get("location", "N/A"),
                profile.get("email", "N/A"),
                profile.get("linkedin", "N/A"),
                ", ".join(profile.get("keywords", [])) if isinstance(profile.get("keywords"), list) else profile.get("keywords", ""),
                profile.get("year", "N/A")
            ]
            ws.append(row_data)
            
            # Format score cell with conditional coloring
            current_row = rank + 1
            score_cell = ws[f"B{current_row}"]
            if profile.get("probability_score", 0) >= 60:
                score_cell.fill = self.score_fill_high
                score_cell.font = self.score_font_high
            
            # Apply borders and alignment to all cells in row
            self._format_data_row(ws, current_row)
        
        # Auto-adjust column widths
        column_widths = {
            'A': 8,    # Rank
            'B': 10,   # Score
            'C': 25,   # Name
            'D': 20,   # Title
            'E': 25,   # Company
            'F': 20,   # Location
            'G': 25,   # Email
            'H': 30,   # LinkedIn
            'I': 40,   # Keywords
            'J': 8,    # Year
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Optionally save to file
        if output_path:
            wb.save(output_path)
        
        return excel_buffer
    
    def _flatten_to_excel(self, data: Dict[str, Any], output_path: str = None) -> BytesIO:
        """
        Convert flattened data to Excel format (original behavior).
        
        Args:
            data: Extracted structured data
            output_path: Optional path to save Excel file
            
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Write header row: #, Key, Value, Comments
        ws.append(["#", "Key", "Value", "Comments"])
        self._format_header_row(ws, 1)
        
        # Flatten data and write rows
        row_number = 1
        current_row = 2
        
        for section_name, section_data in data.items():
            rows = self._flatten_section(section_name, section_data)
            
            for row_data in rows:
                key = row_data.get("key", "")
                value = row_data.get("value", "")
                comments = row_data.get("comments", "")
                
                ws.append([row_number, key, value, comments])
                self._format_data_row(ws, current_row)
                
                row_number += 1
                current_row += 1
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Optionally save to file
        if output_path:
            wb.save(output_path)
        
        return excel_buffer
    
    def _flatten_section(self, section_name: str, section_data: Any) -> List[Dict[str, str]]:
        """
        Flatten a section into a list of rows.
        
        Args:
            section_name: Name of the section
            section_data: Data for the section
            
        Returns:
            List of row dictionaries with key, value, comments
        """
        rows = []
        
        if isinstance(section_data, dict):
            # Simple key-value section
            for key, value in section_data.items():
                if key.lower() == "comments":
                    continue  # Skip standalone comments
                
                # Check if value itself has comments
                comment = ""
                if isinstance(value, dict) and "comments" in value:
                    comment = value.get("comments", "")
                    # If dict has "text" field, use that as the main value
                    if "text" in value:
                        value = value.get("text", "")
                    else:
                        # Remove comments and format remaining dict
                        value = {k: v for k, v in value.items() if k != "comments"}
                
                formatted_value = self._format_value(value)
                
                rows.append({
                    "key": key,
                    "value": formatted_value,
                    "comments": comment
                })
        
        elif isinstance(section_data, list):
            # List of items (e.g., education, work experience)
            for idx, item in enumerate(section_data):
                if isinstance(item, dict):
                    # Extract comments for this item
                    item_comment = item.get("comments", "")
                    
                    # Add all fields from this item
                    first_key = True
                    for key, value in item.items():
                        if key.lower() == "comments":
                            continue
                        
                        formatted_value = self._format_value(value)
                        
                        # Only add comment to the first key of each item
                        rows.append({
                            "key": key,
                            "value": formatted_value,
                            "comments": item_comment if first_key else ""
                        })
                        first_key = False
                else:
                    # Simple list item
                    rows.append({
                        "key": f"item_{idx + 1}",
                        "value": str(item),
                        "comments": ""
                    })
        else:
            # Simple value
            rows.append({
                "key": section_name,
                "value": str(section_data),
                "comments": ""
            })
        
        return rows
    
    def _format_value(self, value: Any) -> str:
        """Format value for Excel cell."""
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        elif isinstance(value, dict):
            # Don't include comments in the formatted value
            filtered = {k: v for k, v in value.items() if k.lower() != "comments"}
            if not filtered:
                return ""
            return "; ".join(f"{k}: {v}" for k, v in filtered.items())
        else:
            return str(value) if value is not None else ""
    
    def _format_header_row(self, ws, row_num: int):
        """Apply formatting to header row."""
        for cell in ws[row_num]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _format_data_row(self, ws, row_num: int):
        """Apply formatting to data row."""
        for idx, cell in enumerate(ws[row_num]):
            cell.border = self.border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Center align the row number column
            if idx == 0:  # First column is row number
                cell.alignment = Alignment(horizontal='center', vertical='top')
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""
        column_widths = {
            'A': 8,   # # (row number)
            'B': 30,  # Key
            'C': 50,  # Value
            'D': 45,  # Comments
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def json_to_dataframe(self, data: Dict[str, Any]) -> pd.DataFrame:
        """
        Convert JSON to pandas DataFrame with flattened structure.
        
        Args:
            data: Extracted structured data
            
        Returns:
            DataFrame representation
        """
        rows = []
        row_number = 1
        
        for section_name, section_data in data.items():
            flattened_rows = self._flatten_section(section_name, section_data)
            
            for row_data in flattened_rows:
                rows.append({
                    "#": row_number,
                    "Key": row_data.get("key", ""),
                    "Value": row_data.get("value", ""),
                    "Comments": row_data.get("comments", "")
                })
                row_number += 1
        
        return pd.DataFrame(rows)